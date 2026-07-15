# ==========================================================
# dashboard_export.py
# InsightOS Dashboard Export Utilities
# ==========================================================

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font


# ==========================================================
# DATASET OVERVIEW
# ==========================================================

def dataset_overview(df):

    return {

        "Rows": len(df),

        "Columns": len(df.columns),

        "Missing Values": int(df.isnull().sum().sum()),

        "Duplicate Rows": int(df.duplicated().sum())

    }


# ==========================================================
# NUMERIC SUMMARY
# ==========================================================

def numeric_summary(df):

    numeric = df.select_dtypes(include="number")

    if numeric.empty:

        return pd.DataFrame()

    summary = pd.DataFrame({

        "Mean": numeric.mean(),

        "Median": numeric.median(),

        "Minimum": numeric.min(),

        "Maximum": numeric.max(),

        "Std Dev": numeric.std()

    })

    return summary.round(2)


# ==========================================================
# CREATE EXCEL DASHBOARD
# ==========================================================

def create_dashboard_workbook(df):

    workbook = Workbook()

    overview_sheet = workbook.active

    overview_sheet.title = "Overview"

    header = Font(

        bold=True,

        size=12

    )

    overview = dataset_overview(df)

    overview_sheet["A1"] = "InsightOS Dashboard Report"

    overview_sheet["A1"].font = Font(

        bold=True,

        size=16

    )

    row = 3

    for key, value in overview.items():

        overview_sheet.cell(

            row=row,

            column=1

        ).value = key

        overview_sheet.cell(

            row=row,

            column=2

        ).value = value

        overview_sheet.cell(

            row=row,

            column=1

        ).font = header

        row += 1

    stats_sheet = workbook.create_sheet(

        "Statistics"

    )

    stats = numeric_summary(df)

    if not stats.empty:

        stats_sheet.append(

            ["Column"] + list(stats.columns)

        )

        for cell in stats_sheet[1]:

            cell.font = header

        for index, values in stats.iterrows():

            stats_sheet.append(

                [index] + list(values)

            )

    preview_sheet = workbook.create_sheet(

        "Dataset Preview"

    )

    preview_sheet.append(

        list(df.columns)

    )

    for cell in preview_sheet[1]:

        cell.font = header

    preview = df.head(100)

    for row in preview.itertuples(index=False):

        preview_sheet.append(

            list(row)

        )

    return workbook


# ==========================================================
# EXPORT DASHBOARD
# ==========================================================

def export_dashboard(df):

    workbook = create_dashboard_workbook(df)

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    return output


# ==========================================================
# EXPORT SUMMARY
# ==========================================================

def export_summary_text(summary):

    lines = []

    for section, value in summary.items():

        lines.append(section.upper())

        lines.append("-" * 40)

        if isinstance(value, dict):

            for k, v in value.items():

                lines.append(f"{k}: {v}")

        elif isinstance(value, list):

            for item in value:

                lines.append(f"• {item}")

        else:

            lines.append(str(value))

        lines.append("")

    return "\n".join(lines)